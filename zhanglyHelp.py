#coding:utf-8

from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import time
reload(sys)
sys.setdefaultencoding('utf8')

#---------计算分年总额------
'''
wb = load_workbook('/Users/SonnyCAO/Desktop/data.xlsx')

sheet = wb.active
print sheet.max_row
print type((sheet.cell(row=4, column=2).value))
for i in range(2,sheet.max_row-1):

    if(sheet.cell(row=i, column=2).value == None):
        sheet.cell(row=i, column=2).value = sheet.cell(row=i-1, column=2).value
    if (sheet.cell(row=i, column=8).value == None):
        sheet.cell(row=i, column=8).value = 0.0

i=2
while(i<sheet.max_row-1):
    print i
    if(sheet.cell(row=i, column=2).value != sheet.cell(row=i+1, column=2).value and sheet.cell(row=i, column=2).value != sheet.cell(row=i-1, column=2).value):
        sheet.cell(row=i, column=10).value = ('%s第此年的总额是%f万元' % ((sheet.cell(row=i, column=2).value),sheet.cell(row=i, column=8).value))
        i=i+1
    else:
        sum = 0.0
        while(str(sheet.cell(row=i, column=7).value)[0:4] == str(sheet.cell(row=i+1, column=7).value)[0:4]):
            sum = sum + float(str(sheet.cell(row=i, column=8).value))
            i = i+1
            if(i>sheet.max_row+1):
                break
        sum =sum + float(str(sheet.cell(row=i, column=8).value))
        sheet.cell(row=i, column=10).value = ('%s第%s年的总额是%f万元' % ((sheet.cell(row=i, column=2).value),str(sheet.cell(row=i, column=7).value)[0:4], sum))

        i = i+1


for i in range(2,sheet.max_row-1):

    if(sheet.cell(row=i, column=8).value == 0.0):
        sheet.cell(row=i, column=8).value = None
    if (sheet.cell(row=i, column=10).value == 0.0):
        sheet.cell(row=i, column=8).value = None

wb.save('/Users/SonnyCAO/Desktop/data_%s.xlsx' % '张')
'''

#-------复制5次

wb1 = Workbook()
sheet1 = wb1.active
wb2 = load_workbook('/Users/SonnyCAO/Desktop/總資產.xlsx')
sheet2 = wb2.active
p=2
for i in range(2,sheet2.max_row+1):
    j=3
    while (j<8):
        sheet1.cell(row=p, column=1).value = i-1
        sheet1.cell(row=p, column=2).value = sheet2.cell(row=i, column=2).value
        sheet1.cell(row=p, column=3).value = sheet2.cell(row=i, column=3).value
        sheet1.cell(row=p, column=4).value = '201%d1231' % j
        j=j+1
        p = p + 1
wb1.save('/Users/SonnyCAO/Desktop/總資產_%s.xlsx' % '张')