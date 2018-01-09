#coding:utf-8
import jieba
from sklearn.cross_validation import train_test_split
from gensim.models.word2vec import Word2Vec
import numpy as np
import pandas as pd
import jieba
from sklearn.externals import joblib
from sklearn.svm import SVC
from openpyxl import load_workbook
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from sklearn.cross_validation import train_test_split
print 'hello'
import sys
reload(sys)
sys.setdefaultencoding('utf8')

neg_wb = load_workbook('/Users/SonnyCAO/Desktop/trainData/Positive_pkubbs.xlsx')
sheet = neg_wb.get_sheet_by_name('q1')

for i in range(2,sheet.max_row+1):
    sheet.cell(row=i, column=3).value = str(sheet.cell(row=i, column=1).value) +'，'+ str(sheet.cell(row=i, column=2).value)

neg_wb.save('/Users/SonnyCAO/Desktop/trainData/Positive_pkubbs.xlsx')
# neg = pd.read_excel('/Users/SonnyCAO/Desktop/trainData/Negtive_pkubbs.xlsx', header=None, index=None)
# pos = pd.read_excel('/Users/SonnyCAO/Desktop/trainData/Positive_pkubbs.xlsx', header=None, index=None)
#
# neg['sentence'] = neg[0] + '-' + neg[6]
# pos['sentence'] = pos[0] + '-' + pos[6]
# for i in range(1,len(neg)):
#     neg['sentence'][i-1] = str(neg['sentence'][i]).replace('\n', '')
#     neg['sentence'][i-1] = str(neg['sentence'][i]).replace(' ', '')
#     neg['sentence'][i-1] = str(neg['sentence'][i]).strip()
#
#     pos['sentence'][i-1] = str(pos['sentence'][i]).replace('\n', '')
#     pos['sentence'][i-1] = str(pos['sentence'][i]).replace(' ', '')
#     pos['sentence'][i-1] = str(pos['sentence'][i]).strip()
#
# # print neg['sentence'][343]
#
#
# cw = lambda x: list(jieba.cut(x))
#
# neg['words'] = neg['sentence'].apply(cw)
# pos['words'] = pos['sentence'].apply(cw)
# #
# #
# # print len(pos['words'])
# print len(neg['words'])

# neg = pd.read_excel('/Users/SonnyCAO/Desktop/trainData/Negtive_pkubbs.xlsx', header=None, index=None)
# pos = pd.read_excel('/Users/SonnyCAO/Desktop/trainData/Positive_pkubbs.xlsx', header=None, index=None)
#
# print (neg[0][900])
# neg['words'] = neg[0]+'，'+neg[6]
# print (neg['words'][900])
# # cw = lambda x: list(x.replace('\n',''))
# # print neg['words'].apply(cw)
# for i in range(len(neg)):
#     neg['words'][i] = str(neg['words'][i]).replace('\n','')
#     neg['words'][i] = str(neg['words'][i]).replace(' ', '')
#     neg['words'][i] = str(neg['words'][i]).strip()
#
# print (neg['words'][900])



# cw = lambda x: list(jieba.cut(x))
# pos['words'] = pos[0].apply(cw)
# neg['words'] = neg[0].apply(cw)
#
# # print pos['words']
# # use 1 for positive sentiment, 0 for negative
# y = np.concatenate((np.ones(len(pos)), np.zeros(len(neg))))
#
# x_train, x_test, y_train, y_test = train_test_split(np.concatenate((pos['words'], neg['words'])), y, test_size=0.2)
#
# np.save('svm_data/y_train.npy', y_train)
# np.save('svm_data/y_test.npy', y_test)


# def readLines(filename):
#     fopen = open(filename, 'r')
#     data=[]
#     for x in fopen.readlines():
#         if x.strip() != '':
#             data.append(x.strip())
#     fopen.close()
#     return data
#
# def dataPreProcess():
#     neg_wb = load_workbook('/Users/SonnyCAO/Desktop/trainData/Negtive_pkubbs.xlsx')
#     pos_wb = load_workbook('/Users/SonnyCAO/Desktop/trainData/Positive_pkubbs.xlsx')
#
#     neg_sheet = neg_wb.active
#     pos_sheet = pos_wb.active
#
#     neg_words = []
#     for i in range(2,neg_sheet.max_row+1):
#         p1 = neg_sheet.cell(row=i, column=1).value
#         p2 = neg_sheet.cell(row=i, column=7).value
#         post = str(p1) + '，' + str(p2)
#         post = post.replace('\n', '')
#         post = post.replace(' ', '')
#         post = post.strip()
#         neg_words.append(wordCutAndRemoveStopWords(post))
#
#     pos_words = []
#     for i in range(2, pos_sheet.max_row+1):
#         p1 = pos_sheet.cell(row=i, column=1).value
#         p2 = pos_sheet.cell(row=i, column=7).value
#         post = str(p1) + '，' + str(p2)
#         post = post.replace('\n', '')
#         post = post.replace(' ', '')
#         post = post.strip()
#         pos_words.append(wordCutAndRemoveStopWords(post))
#
#     y = np.concatenate((np.ones(len(pos_words)), np.zeros(len(neg_words))))
#     x_train, x_test, y_train, y_test = train_test_split(np.concatenate((pos_words, neg_words)), y, test_size=0.2)
#
#     np.save('svm_data/y_train', y_train)
#     np.save('svm_data/y_test', y_test)
#     return x_train, x_test
#
#
# def wordCutAndRemoveStopWords(sentence):
#     #结巴分词
#     segResult = jieba.lcut(sentence)
#
#     #去掉停用词
#     stopwords = readLines('/Users/SonnyCAO/Desktop/trainData/stopwords.txt')
#     newSent = []
#     for word in segResult:
#         if word in stopwords:
#             continue
#         else:
#             newSent.append(word)
#     return newSent
#
# # string = ' 我好迷茫，令人痛苦的心理困扰'
# # words = jieba.lcut(string)
# # print type(words)
# # for word in words:
# #     print (word)
# # print np.zeros(300)
# a = np.zeros(300).reshape((1, 300))
# # print (a)
# # print a
# model = Word2Vec.load('svm_data/w2v_model.pkl')
# print model[u'死']
# print model[u'死'].reshape((1,300))
# print a+model[u'死'].reshape((1,300))
#
# # neg = load_workbook('/Users/SonnyCAO/Desktop/trainData/Negtive_pkubbs.xlsx')
# # sheet = neg.active
# # n = neg.get_sheet_names()
# # # print type(n)
# # shet = neg.get_sheet_by_name(n[0])
# # print sheet.max_row
# # p1 = sheet.cell(row=1,column=1).value
# # p2 = sheet.cell(row=1,column=7).value
# # post = str(p1)+'，'+str(p2)
# # print post
# # print type(post)
# #
# # # post = (post.encode('utf-8'))
# # post = post.replace('\n','')
# # post = post.replace(' ','')
# # post = post.strip()
# # print post
# # segResult = []
# #
# # segList = jieba.cut(post)
# # print type(segList)
# #
# # for w in segList:
# #     segResult.append(w)
# #
# # print len(segResult)
# # y = np.concatenate((np.ones(sheet.max_row), np.zeros(sheet.max_row)))
# # w = np.concatenate((np.ones(2), np.zeros(3)))
# # print len(np.ones(sheet.max_row))
# # print w
# #
# # # # for item in segResult:
# # # #     print item
# # stopwords = readLines('/Users/SonnyCAO/Desktop/trainData/stopwords.txt')
# # newSent=[]
# # stopwords_list=[]
# # for word in segResult:
# #     if word in stopwords:
# #         # print "stopword: %s" % word
# #         continue
# #     else:
# #         newSent.append(word)
# #         print word
# # #
