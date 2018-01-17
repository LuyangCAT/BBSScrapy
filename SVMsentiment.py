# -*- coding: utf-8 -*-
"""
@author: cly
"""
from sklearn.cross_validation import train_test_split
from gensim.models.word2vec import Word2Vec
import numpy as np
import pandas as pd
import jieba
from sklearn.preprocessing import scale
from sklearn.externals import joblib
from sklearn.svm import SVC
from sklearn.linear_model import LogisticRegression
from openpyxl import load_workbook
from sklearn.decomposition import PCA

from sklearn.ensemble import AdaBoostClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.linear_model import Perceptron
from sklearn.linear_model import SGDClassifier
from sklearn.model_selection import KFold, cross_val_score, train_test_split
from sklearn.model_selection import learning_curve
from sklearn.naive_bayes import GaussianNB
from sklearn.neighbors import KNeighborsClassifier
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.svm import SVC, LinearSVC
from sklearn.tree import DecisionTreeClassifier

import matplotlib.pyplot as plt
from sklearn.metrics import roc_curve, auc
import sys

reload(sys)
sys.setdefaultencoding('utf8')

#读取文件
def readLines(filename):
    fopen = open(filename, 'r')
    data=[]
    for x in fopen.readlines():
        if x.strip() != '':
            data.append(x.strip())
    fopen.close()
    return data
#数据预处理
def dataPreProcess():
    neg_wb = load_workbook('/Users/SonnyCAO/Desktop/trainData/neg_final.xlsx')
    pos_wb = load_workbook('/Users/SonnyCAO/Desktop/trainData/pos_final.xlsx')

    neg_sheet = neg_wb.active
    pos_sheet = pos_wb.active

    neg_words = []
    for i in range(1,neg_sheet.max_row+1):
        print i
        post = neg_sheet.cell(row=i, column=1).value
        # p2 = neg_sheet.cell(row=i, column=7).value
        # post = str(p1) + '，' + str(p2)
        post = str(post).replace('\n', '')
        post = post.replace(' ', '')
        post = post.strip()
        neg_words.append(wordCutAndRemoveStopWords(post))

    pos_words = []
    for i in range(2, pos_sheet.max_row+1):
        print i
        post = pos_sheet.cell(row=i, column=1).value
        # p2 = pos_sheet.cell(row=i, column=7).value
        # post = str(p1) + '，' + str(p2)
        post = str(post).replace('\n', '')
        post = post.replace(' ', '')
        post = post.strip()
        pos_words.append(wordCutAndRemoveStopWords(post))

    y = np.concatenate((np.ones(len(pos_words)), np.zeros(len(neg_words))))
    # X = X_pca(pos_words, neg_words)
    x_train, x_test, y_train, y_test = train_test_split(np.concatenate((pos_words,neg_words)), y, test_size=0.2)

    np.save('svm_data/x_train', x_train)
    np.save('svm_data/x_test', x_test)
    np.save('svm_data/y_train', y_train)
    np.save('svm_data/y_test', y_test)
    return x_train, x_test


def wordCutAndRemoveStopWords(sentence):
    #结巴分词
    segResult = []
    segList = jieba.cut(sentence)
    for w in segList:
        segResult.append(w)
    #去掉停用词-`                                                                                (1)
    # stopwords = readLines('/Users/SonnyCAO/Desktop/trainData/stopwords.txt')
    # newSent = []
    # for word in segResult:
    #     if word in stopwords:
    #         continue
    #     else:
    #         newSent.append(word)
    # return newSent
    return segResult


# 对每个句子的所有词向量取均值
def buildWordVector(text, size, model_w2v):
    vec = np.zeros(size).reshape((1, size))
    count = 0.
    for word in text:

        try:
            vec += model_w2v[word].reshape((1, size))
            count += 1.
        except KeyError:
            continue
    # print vec--                                                                               (2)归一化
    if count != 0:
        vec /= count
    # print vec
    # vec  = scale(vec)
    return vec


# 计算词向量
def get_train_vecs(x_train, x_test):
    n_dim = 300
    model_w2v = Word2Vec(x_train,min_count=3,size=300)



    # # Initialize model and build vocab
    # model_w2v = Word2Vec(size=n_dim, min_count=1)---                                              (3)
    # model_w2v.build_vocab(x_train)
    #
    # # Train the model over train_reviews (this may take several minutes)
    # model_w2v.train(x_train,total_examples=model_w2v.corpus_count,epochs=1)
    #
    train_vecs = np.concatenate([buildWordVector(z, n_dim, model_w2v) for z in x_train])
    # train_vecs = scale(train_vecs)
    #
    np.save('svm_data/train_vecs', train_vecs)
    print train_vecs.shape
    # # Train word2vec on test tweets
    # model_w2v.train(x_test,total_examples=model_w2v.corpus_count,epochs=1)
    model_w2v.save('svm_data/w2v_model.pkl')
    # # Build test tweet vectors then scale
    test_vecs = np.concatenate([buildWordVector(z, n_dim, model_w2v) for z in x_test])
    # # test_vecs = scale(test_vecs)
    np.save('svm_data/test_vecs', test_vecs)
    print test_vecs.shape


def get_data():
    train_vecs = np.load('svm_data/train_vecs.npy')
    y_train = np.load('svm_data/y_train.npy')
    test_vecs = np.load('svm_data/test_vecs.npy')
    y_test = np.load('svm_data/y_test.npy')
    return train_vecs, y_train, test_vecs, y_test


##训练svm模型
def svm_train(train_vecs, y_train, test_vecs, y_test):
    clf = SVC(kernel='rbf',  verbose=True)#------------------------（4）换成线性拟合效果变成93%,径向基内核80%，
    clf.fit(train_vecs, y_train)
    joblib.dump(clf, 'svm_data/SVC-rbf_model.pkl')
    print clf.score(test_vecs, y_test),'SVC-rbf'

    clf = LinearSVC()  # ------------------------（4）换成线性拟合效果变成93%,径向基内核80%，
    clf.fit(train_vecs, y_train)
    joblib.dump(clf, 'svm_data/SVC-linear_model.pkl')
    print clf.score(test_vecs, y_test), 'SVC-linear'

    clf1 = LogisticRegression()
    clf1.fit(train_vecs, y_train)
    joblib.dump(clf1, 'svm_data/logistc_model.pkl')
    print clf1.score(test_vecs, y_test), 'Logistic'

    clf1 = AdaBoostClassifier()
    clf1.fit(train_vecs, y_train)
    joblib.dump(clf1, 'svm_data/AdaBoostClassifier_model.pkl')
    print clf1.score(test_vecs, y_test), 'AdaBoostClassifier'

    clf1 = RandomForestClassifier()
    clf1.fit(train_vecs, y_train)
    joblib.dump(clf1, 'svm_data/RandomForestClassifier_model.pkl')
    print clf1.score(test_vecs, y_test), 'RandomForestClassifier'

    clf1 = KNeighborsClassifier()
    clf1.fit(train_vecs, y_train)
    joblib.dump(clf1, 'svm_data/lKNeighborsClassifier_model.pkl')
    print clf1.score(test_vecs, y_test), 'KNeighborsClassifier'

    clf1 = GaussianNB()
    clf1.fit(train_vecs, y_train)
    joblib.dump(clf1, 'svm_data/GaussianNB_model.pkl')
    print clf1.score(test_vecs, y_test), 'GaussianNB'

    clf1 = SGDClassifier()
    clf1.fit(train_vecs, y_train)
    joblib.dump(clf1, 'svm_data/SGDClassifier_model.pkl')
    print clf1.score(test_vecs, y_test), 'SGDClassifier'

    clf1 = DecisionTreeClassifier()
    clf1.fit(train_vecs, y_train)
    joblib.dump(clf1, 'svm_data/DecisionTreeClassifier_model.pkl')
    print clf1.score(test_vecs, y_test), 'DecisionTreeClassifier'

    clf1 = Perceptron()
    clf1.fit(train_vecs, y_train)
    joblib.dump(clf1, 'svm_data/Perceptron_model.pkl')
    print clf1.score(test_vecs, y_test), 'Perceptron'









##得到待预测单个句子的词向量
def get_predict_vecs(words):
    n_dim = 300
    imdb_w2v = Word2Vec.load('svm_data/w2v_model.pkl')
    # imdb_w2v.train(words)
    train_vecs = buildWordVector(words, n_dim, imdb_w2v)
    # print train_vecs.shape
    return train_vecs


####对单个句子进行情感判断
def svm_predict(string):
    words = wordCutAndRemoveStopWords(string)
    # for word in words:
    #     print word
    words_vecs = get_predict_vecs(words)
    # print words_vecs
    clf = joblib.load('svm_data/RandomForestClassifier_model.pkl')

    result = clf.predict(words_vecs)
    # print type(result),len(result),result[0]

    if int(result[0]) == 1:
        print string
        print ' positive'
    else:
        print string
        print ' negative'

def X_pca(posInput,negInput):

    # pca.fit(X)
    # plt.figure(1, figsize=(4, 3))
    # plt.clf()
    # plt.axes([.2, .2, .7, .7])
    # plt.plot(pca.explained_variance_, linewidth=2)
    # plt.axis('tight')
    # plt.xlabel('n_components')
    # plt.ylabel('explained_variance_')
    X = posInput[:]
    for neg in negInput:
        X.append(neg)
    X = np.array(X)
    X = scale(X)
    X_reduced = PCA(n_components=100).fit_transform(X)
    return X_reduced


if __name__ == '__main__':
    #导入文件，处理保存为向量
    # x_train,x_test=dataPreProcess() #得到句子分词后的结果，并把类别标签保存为y_train。npy,y_test.npy
    # x_train = np.load('svm_data/x_train.npy')
    # x_test = np.load('svm_data/x_test.npy')
    # # for it in x_train:
    # #     for i in range(len(it)):
    # #         print it[i]


    # x_train,x_test=dataPreProcess()
    # get_train_vecs(x_train,x_test) #计算词向量并保存为train_vecs.npy,test_vecs.npy
    # train_vecs,y_train,test_vecs,y_test=get_data()#导入训练数据和测试数据
    # svm_train(train_vecs,y_train,test_vecs,y_test)#训练svm并保存模型

    # clf = joblib.load('svm_data/svm_model.pkl')
    # print clf.score(test_vecs, y_test)
    ##对输入句子情感进行判断
    string = u'不开心不开心不开心不开心不开心不开心不开心'  #-                         ？这是个问题？
    string1 = u'很开心'
    svm_predict(string)
    svm_predict(string1)
    # for i in range(2,20):
    #     neg_wb = load_workbook('/Users/SonnyCAO/Desktop/trainData/Negtive_pkubbs.xlsx')
    #     neg_sheet = neg_wb.active
    #     p1 = neg_sheet.cell(row=i, column=1).value
    #     p2 = neg_sheet.cell(row=i, column=7).value
    #     post = str(p1) + '，' + str(p2)
    #     post = post.replace('\n', '')
    #     post = post.replace(' ', '')
    #     post = post.strip()
    #     svm_predict(post)

#1词向量到句向量这块的算法
#2分类的模型换一换以及模型的参数


