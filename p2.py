#coding:utf-8

import urllib2
import re
import xlwt
import sys
import time
import requests
from openpyxl import Workbook

reload(sys)
sys.setdefaultencoding('utf8')


class BBSSpider:


    def __init__(self):
        self.baseURL = ""
        self.enable = True
        self.charaterset = "gb2312"

    # 发送URL请求获得页面信息
    def getHtml(self, url):

        #模拟登陆
        # user_agent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_2) ' \
        #              'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.84 Safari/537.36'
        # referer = 'https://bbs.pku.edu.cn/v2/login.php'
        #
        # headers = {
        #     'User-Agent': user_agent,
        #     'Host': 'bbs.pku.edu.cn',
        #     'Origin': 'https://bbs.pku.edu.cn/v2/login.php',
        #     'Referer': referer
        # }
        #
        # cookie = {'cookie':'uid=12694; skey=602fb6e6d3020a04'}
        # htt = requests.get(url,cookies=cookie,headers=headers)
        # print htt.content
        #

        self.baseURL = url
        try:
            request = urllib2.Request(self.baseURL)
            response = urllib2.urlopen(request)
            return response.read().decode("utf-8", 'ignore')
        except urllib2.URLError, e:
            if hasattr(e, "reason"):
                string = "连接bbs失败, 原因：" +  str(e.reason)
                print string.encode(self.charaterset)
                return None

    # 处理合集后的网页
    def processHeJi(self, content):
        # pp = r'<div class="post-card".*?>.*?<p class="username">.*?<a .*?>' \
        #      r'(.*?)</a>.*?<div class="body file-read image-click-view">(.*?)' \
        #      r'(<p class="quotehead".*?>|</div>|<p class="blockquote".*?>).*?<div class="sl-triangle-container"' \
        #      r'.*?>.*?<span>\D{3,10}(.*?)</span>.*?</div>'
        removeSp = re.compile(r'<font.*?>|</font>|<span style="background-color.*?>|</span>|<p .*?>|</p>|<a .*?>|</a>')
        content = re.sub(removeSp, " ", content).strip()
        pat = ur'<div class="body file-read image-click-view".*?作者(.*?)\(.*?时间.*?\((.*?)星期.*?───────────────────────────────────────(.*?)───────────────────────────────────────.*?</div>'
        patt = re.compile(pat, re.S)
        res = re.findall(patt, content)

        if len(res) == 0:
            return None
        posts = []
        for i in range(len(res)):
            col = [[],[],[],[]]
            col[0] = res[i][0].strip()
            col[1] = res[i][2].strip()
            col[2] = res[i][0].strip()
            col[3] = res[i][1].strip()
            posts.append(col)
        #print len(posts),posts[0][1]
        return posts

    # 删除网页内容中的一些噪声
    def removeNoise(self, content):

        removeFont = re.compile(r'<font.*?>|</font>|<span style="background-color.*?>')
        content = re.sub(removeFont, " ", content).strip()
        removeNBSP = re.compile(r"&nbsp;|<b>|</b>|&gt;|&lt;|&quot;")
        content = re.sub(removeNBSP, " ", content).strip()
        removeAMP = re.compile(r"&amp;")
        content = re.sub(removeAMP, "&", content).strip()
        removeBR = re.compile(r"<br/>")
        content = re.sub(removeBR, "\n", content).strip()
        removeP = re.compile(r'<tr>|<p>|</p>')
        content = re.sub(removeP, "\n", content).strip()
        removeN = re.compile(r"\n{1,}")
        content = re.sub(removeN, "\n", content).strip()

        return content

    # 获取发帖信息
    def getItem(self, board):
        print '正在获取该版面的总页数...'

        string = "https://bbs.pku.edu.cn/v2/thread.php?bid=" + str(board)
        content = self.getHtml(string)
        if not content:
            print "加载页面失败"
            return

        p = r'<div>/(.*?)</div>'  #获取页面总数的正则表达式
        pattern = re.compile(p, re.S)
        pages = re.findall(pattern, content)

        print '本版面总共有%d页帖子...' %(int(pages[0]))

        titles = []
        for i in range(int(pages[0])):#暂时只用2页   int(pages[0])
            url = string + "&mode=topic&page=" + str(i+1)
            content = self.getHtml(url)
            content = self.removeNoise(content)

            # 获取标题的正则表达式，格式[链接][题目][作者][时间]
            pt = r'<div class="list-item-topic list-item" .*?><.*? href="(.*?)">' \
                 r'</a>.*?<div class="title l limit".*?>(.*?)</div>.*?<div class="au' \
                 r'thor l">.*?<div class="name limit">(.*?)</div>.*?<div class="time">(.*?)</div>'
            pat = re.compile(pt,re.S)
            res = re.findall(pat,content)

            spList = [ u"Re:", u"招募", u'转载',u'公告',u'诚招',u'提示',u'投票',u'招新',u'访谈',u'讲座',u'征集',u'本期',u'活动',u'实验']  #对于含有此类字符的标题不做爬取
            autList = [u'weiqj', u'lianggq', u'wangss',u'stonerex', u'miaoga', u'ding', u'hibisay',u'kidkidkid',u'yuyin',u'deliver',u'lisongwei']
            f1 = False  #标记标题中是否出现特殊字段
            f2 = False  #标记作者是否为管理员
            for item in res:
                for it in spList:
                    if it in item[1]:
                        f1 = True
                        break
                for itt in autList:
                    if itt in item[1]:
                        f2 = True
                        break
                if f1 or f2:
                    continue
                # if u"[合集]" in item[1]:   #标题出现合集就跳过
                #     continue
                # elif u"Re:" in item[1]:
                #     continue
                # elif u"【招募】" in item[1]:
                #     continue
                # elif u'转载' in item[1]:
                #     continue
                # elif item[2] == u"deliver" or item[2] == u"weiqj" or \
                #                 item[2] == u"lianggq" or item[2] == u"wangss"or \
                #                 item[2] == u"kidkidkid" or item[2] == u'stonerex'\
                #         or item[2] == u'hibisay' or item[2] == u'miaoga' or item[2] == u'ding':
                #     continue
                else:
                    #print item[1]
                    titles.append(item)
                    #print item[1],item[2]
        print '本版面总共有%d个有效标题...' % (len(titles))
        return titles


    # 获取详细信息
    def getDetails(self, id,name):
        titles = self.getItem(id)
        if not titles:
            return

        time.sleep(1)

        #创建Excel文件
        wb = Workbook()
        sheet = wb.active
        sheet.title = name
        heads = [u'话题', u'话题作者id', u'话题开启时间', u'发言人id', u'发言时间',u'样本极性', u'发言内容']  # 写入字段名称
        ii = 1
        for head in heads:
            sheet.cell(row=1, column=ii).value = head
            ii += 1




        #
        #
        # workbook = xlwt.Workbook(encoding='utf-8')
        # worksheet = workbook.add_sheet(name,cell_overwrite_ok=True)
        # heads = [u'话题', u'话题作者id', u'话题开启时间', u'发言人id', u'发言时间', u'发言内容'] #写入字段名称
        # ii = 0
        # for head in heads:
        #     worksheet.write(0, ii, head)
        #     ii += 1

        ii = 2                #控制写入行的计数
        count = 1            #控制命令行提示数字的计数
        for title in titles:
            url = 'https://bbs.pku.edu.cn/v2/' + str(title[0])
            # content = self.getHtml(url)
            # content = self.removeNoise(content)
            # p = r'<div>/(.*?)</div>'
            # pattern = re.compile(p, re.S)
            # pages = re.findall(pattern, content)

            print '正在写入Excel文档中,请稍后...%d' %count
            # for page in range(2): #只在每个帖子前两页搜索咨询师的id     int(pages[0])
            url1 = url + '&page='+str(1)
            content = self.getHtml(url1)
            content = self.removeNoise(content)


            flag = False  # 判断有没有心理老师的回复
            Psychological_Consultant_list = [u'weiqj', u'lianggq', u'wangss',
                                             u'stonerex', u'miaoga', u'ding',
                                             u'hibisay', u'kidkidkid', u'yuyin',
                                             u'lisongwei', u'qinmo', u'niejing']

            #合集处理办法
            if u'合集' in title[1]:
                posts = self.processHeJi(content)
                #print posts[0][1],len(posts[0])
            else:

                # 获取发言内容的正则表达式，格式[作者][发言内容]【废】[时间]
                pp = r'<div class="post-card".*?>.*?<p class="username">.*?<a .*?>' \
                     r'(.*?)</a>.*?<div class="body file-read image-click-view">(.*?)' \
                     r'(<p class="quotehead".*?>|</div>|<p class="blockquote".*?>).*?<div class="sl-triangle-container"' \
                     r'.*?>.*?<span>\D{3,10}(.*?)</span>.*?</div>'
                patt = re.compile(pp, re.S)
                posts = re.findall(patt, content)
                for teacher in Psychological_Consultant_list:
                    for post in posts:
                        if teacher in post[0]:
                            flag = True
                            break

            if posts is None:
                continue





            #写入Excel
            ppp = posts[0][1]
            removeSp = re.compile(r'<font.*?>|</font>|<span style="background-color.*?>|</span>|<p .*?>|</p>|<a .*?>|</a>')
            conn = re.sub(removeSp, " ", ppp).strip()
            try:
                sheet.cell(row=ii, column=7).value = conn

            except:
                print "----特殊字符串Exception----"
                continue
            else:
                sheet.cell(row=ii, column=1).value = title[1]
                sheet.cell(row=ii, column=2).value = title[2]
                sheet.cell(row=ii, column=3).value = title[3]
                sheet.cell(row=ii, column=4).value = posts[0][0]
                sheet.cell(row=ii, column=5).value = posts[0][3]
                if flag is True or u'合集' in title[1]: #有老师回复的帖子和合集的帖子标记样本为负样本，记号为'0'
                    sheet.cell(row=ii, column=6).value = '0'
                ii += 1
                # for post in posts:
                #     # if post[1][2:4] == '[I':
                #     #     continue
                #     # post[1] = self.removeNoise(post[1])
                #     #print post[1][2:4]
                #     ppp = post[1]
                #     removeSp = re.compile(r'<font.*?>|</font>|<span style="background-color.*?>|</span>|<p .*?>|</p>|<a .*?>|</a>')
                #     conn = re.sub(removeSp, " ", ppp).strip()
                #     try:
                #         sheet.cell(row=ii, column=6).value = conn
                #
                #     except:
                #         print "----特殊字符串Exception----"
                #         continue
                #     else:
                #         sheet.cell(row=ii, column=1).value = title[1]
                #         sheet.cell(row=ii, column=2).value = title[2]
                #         sheet.cell(row=ii, column=3).value = title[3]
                #         sheet.cell(row=ii, column=4).value = post[0]
                #         sheet.cell(row=ii, column=5).value = post[3]
                #         wb.save('/Users/SonnyCAO/Desktop/pkuBBSm.xlsx')
                #         # worksheet.write(ii, 0, title[1])
                #         # worksheet.write(ii, 1, title[2])
                #         # worksheet.write(ii, 2, title[3])
                #         # worksheet.write(ii, 3, post[0])
                #         # worksheet.write(ii, 4, post[3])
                #         # worksheet.write(ii, 5, post[1])
                #         # workbook.save('/Users/SonnyCAO/Desktop/bbs.xls')
                #         ii += 1
            count += 1
            wb.save('/Users/SonnyCAO/Desktop/pkuBBSm.xlsx')
        print "打印完成~"


    # 获得板块信息
    def getBoard(self):

        url = "https://bbs.pku.edu.cn/v2/board.php?bid=679"
        request = urllib2.Request(url)
        response = urllib2.urlopen(request)
        content = response.read().decode('utf-8')

        pattern = re.compile(r'<div class="board-block" data-bid="(.*?)">.*?<span class="name">(.*?)</span>.*?</div>', re.S)
        boards = re.findall(pattern, content)

        boards = set(boards)

        for b in boards:
            print b[1],b[0]

        string_tip = str("\n请输入要爬取的版面的id号码\n")
        id = raw_input(string_tip)
        for i in boards:
            if id == i[0]:
                self.getDetails(id,i[1])



if "__main__" == __name__:
    bbs = BBSSpider()
    #bbs.getItem(690)
    bbs.getDetails(690, u'心理健康教育与咨询中心')

