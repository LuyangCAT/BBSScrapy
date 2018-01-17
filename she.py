#coding:utf-8

from openpyxl import Workbook
from openpyxl import load_workbook
import sys




wbDes = load_workbook('/Users/SonnyCAO/Desktop/qqqqqq.xlsx')
sheet1 = wbDes.active

wbSrc  = load_workbook('/Users/SonnyCAO/Desktop/wwwwwww.xlsx')
sheet2 = wbSrc.get_sheet_by_name('zly')

i=2
q=2
while(i<sheet1.max_row+1)   :
    for j in range(5,0,-1):
        sheet1.cell(row=i,column=10).value = sheet2.cell(row=q,column=j).value
        i+=1
    q+=1

wbDes.save('/Users/SonnyCAO/Desktop/qqqqqq.xlsx')
wbSrc.save('/Users/SonnyCAO/Desktop/wwwwwww.xlsx')





